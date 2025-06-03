import pandas as pd
from pathlib import Path
from rich import print

import openpyxl
from openpyxl.utils import get_column_letter
import openpyxl.styles
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Border, Side, PatternFill, Font, NamedStyle, GradientFill, Alignment

# Define some openpyxl styles for the report
STYLE_DEFINITIONS = {
    'GoodBadNA': {
        'fill': PatternFill(start_color='63EF45', end_color='63EF45', fill_type='solid'),
        'font': Font(color='63EF45')
    },
    'GoodBad': {
        'fill': PatternFill(start_color='63EF45', end_color='63EF45', fill_type='solid'),
        'font': Font(color='63EF45')
    },
    'Good': {
        'fill': PatternFill(start_color='63EF45', end_color='63EF45', fill_type='solid'),
        'font': Font(color='63EF45')
    },
    'Bad': {
        'fill': PatternFill(start_color='C0504D', end_color='C0504D', fill_type='solid'),
        'font': Font(color='C0504D')
    },
    'XBlank': {
        'fill': PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid'),
        'font': Font(color='FFD966')
    },
    'Navigation': {
        'font': Font(color='0000AA', size=14)
    },
    'Header': {
        'font': Font(bold=True, size=11),
        'fill': PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid'),
        'border': Border(left=Side(style='thick', color='000000'),
                        right=Side(style='thick', color='000000'),
                        top=Side(style='thick', color='000000'),
                        bottom=Side(style='thick', color='000000'))
    },
    'TrafficLight': {
        'font': Font(color='000000'),
        'fill': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'),
        'border': Border(left=Side(style='thick', color='000000'),
                        right=Side(style='thick', color='000000'),
                        top=Side(style='thick', color='000000'),
                        bottom=Side(style='thick', color='000000'))
    },
    'Information': {
        'font': Font(color='000000', size=9, italic=True),
        'fill': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    }
}

def create_style_guide():
    """Create an Excel workbook showing all available styles and conditional formatting."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Style Guide"
    
    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 40
    
    # Add header
    ws['A1'] = "Style Name"
    ws['B1'] = "Example"
    ws['C1'] = "Description"
    ws.row_dimensions[1].height = 30
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    
    row = 2
    
    # Direct styles from STYLE_DEFINITIONS
    for style_name, style_def in STYLE_DEFINITIONS.items():
        ws[f'A{row}'] = style_name
        ws[f'B{row}'] = "Sample Text"
        ws[f'C{row}'] = f"Direct style application"
        
        cell = ws[f'B{row}']
        if 'font' in style_def:
            cell.font = style_def['font']
        if 'fill' in style_def:
            cell.fill = style_def['fill']
        if 'border' in style_def:
            cell.border = style_def['border']
            
        row += 1
    
    # Add spacing
    row += 1
    
    # Conditional formatting examples
    ws[f'A{row}'] = "Conditional Formatting"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    
    # ZeroOne
    ws[f'A{row}'] = "ZeroOne"
    ws[f'B{row}'] = "0"
    ws[f'B{row+1}'] = "1"
    ws[f'C{row}'] = "0 = White, 1 = Light Orange"
    range_ref = f"B{row}:B{row+1}"
    ws.conditional_formatting.add(range_ref,
        CellIsRule(operator='equal', formula=['0'], 
                  fill=PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')))
    ws.conditional_formatting.add(range_ref,
        CellIsRule(operator='equal', formula=['1'], 
                  fill=PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')))
    row += 2

    # ZeroTwo  
    ws[f'A{row}'] = "ZeroTwo"
    ws[f'B{row}'] = "0"
    ws[f'B{row+1}'] = "2"
    ws[f'C{row}'] = "0 = Light Red, 2 = Light Green"
    range_ref = f"B{row}:B{row+1}"
    ws.conditional_formatting.add(range_ref,
        CellIsRule(operator='equal', formula=['0'],
                  fill=PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')))
    ws.conditional_formatting.add(range_ref,
        CellIsRule(operator='equal', formula=['2'],
                  fill=PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid')))
    row += 2

    # TrueFalse
    ws[f'A{row}'] = "TrueFalse"
    ws[f'B{row}'] = "TRUE"
    ws[f'B{row+1}'] = "FALSE"
    ws[f'C{row}'] = "TRUE = Light Green, FALSE = Light Red"
    range_ref = f"B{row}:B{row+1}"
    ws.conditional_formatting.add(range_ref,
        CellIsRule(operator='equal', formula=['TRUE'],
                  fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')))
    ws.conditional_formatting.add(range_ref,
        CellIsRule(operator='equal', formula=['FALSE'],
                  fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')))
    row += 2

    # GoodBadNA
    ws[f'A{row}'] = "GoodBadNA"
    ws[f'B{row}'] = "1"
    ws[f'B{row+1}'] = "0"
    ws[f'B{row+2}'] = ""
    ws[f'C{row}'] = "1 = Green, 0 = Red, Blank = White"
    range_ref = f"B{row}:B{row+2}"
    ws.conditional_formatting.add(range_ref,
        CellIsRule(operator='equal', formula=['1'],
                  fill=PatternFill(start_color='63EF45', end_color='63EF45', fill_type='solid'),
                  font=Font(color='63EF45')))
    ws.conditional_formatting.add(range_ref,
        CellIsRule(operator='equal', formula=['0'],
                  fill=PatternFill(start_color='C0504D', end_color='C0504D', fill_type='solid'),
                  font=Font(color='C0504D')))
    ws.conditional_formatting.add(range_ref,
        CellIsRule(operator='equal', formula=['""'],
                  fill=PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'),
                  font=Font(color='FFFFFF')))
    row += 3

    # XBlank
    ws[f'A{row}'] = "XBlank"
    ws[f'B{row}'] = "X"
    ws[f'B{row+1}'] = ""
    ws[f'C{row}'] = "X = Light Orange, Blank = White"
    range_ref = f"B{row}:B{row+1}"
    ws.conditional_formatting.add(range_ref,
        FormulaRule(formula=[f'EXACT(B{row},"X")'],
                    fill=PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')))
    ws.conditional_formatting.add(range_ref,
        FormulaRule(formula=[f'ISBLANK(B{row})'],
                    fill=PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')))
    row += 2

    # Font styles
    ws[f'A{row}'] = "Bold"
    ws[f'B{row}'] = "Sample Text"
    ws[f'C{row}'] = "Bold text when not empty"
    range_ref = f"B{row}"
    ws.conditional_formatting.add(range_ref,
        FormulaRule(formula=[f'LEN(TRIM(B{row}))>0'],
                    font=Font(bold=True)))
    row += 1

    ws[f'A{row}'] = "Italic"
    ws[f'B{row}'] = "Sample Text"
    ws[f'C{row}'] = "Italic text when not empty"
    range_ref = f"B{row}"
    ws.conditional_formatting.add(range_ref,
        FormulaRule(formula=[f'LEN(TRIM(B{row}))>0'],
                    font=Font(italic=True)))
    # save the workbook
    wb.save('style_guide.xlsx')
    return

def setup_navigation_row(ws, navigation_row):
    for cell in ws[navigation_row]:
        cell.style = 'navigation'

    # Insert a Hyperlink to the tab "Overview" in column A
    ws.cell(row=navigation_row, column=1).value = "Go to Overview"
    ws.cell(row=navigation_row, column=1).hyperlink = "Overview!A1"

def apply_header_style(ws, header_row):
    for cell in ws[header_row]:
        cell.style = 'header'

def applyFontStyleToCell(cell, style):
    if style is None:
        return
    if style == 'Italic':
        cell.font = openpyxl.styles.Font(italic=True)
    elif style == 'Bold':
        cell.font = openpyxl.styles.Font(bold=True)
    elif style == 'Underline':
        cell.font = openpyxl.styles.Font(underline='single')
    elif style == 'Strike':
        cell.font = openpyxl.styles.Font(strike=True)
    elif style == 'ItalicBold':
        cell.font = openpyxl.styles.Font(italic=True, bold=True)
    elif style == 'ItalicUnderline':
        cell.font = openpyxl.styles.Font(italic=True, underline='single')
    elif style == 'ItalicStrike':
        cell.font = openpyxl.styles.Font(italic=True, strike=True)
    elif style == 'BoldUnderline':
        cell.font = openpyxl.styles.Font(bold=True, underline='single')
    elif style == 'BoldStrike':
        cell.font = openpyxl.styles.Font(bold=True, strike=True)

def applyFontSizeToCell(cell, size):
    if size is None or pd.isna(size):
        return
    cell.font = openpyxl.styles.Font(size=size)

def applyFontColorToCell(cell, color):
    if color is None or pd.isna(color):
        return
    # Convert color name to hex if needed
    if isinstance(color, (str, float)):
        # Convert float to hex string if needed
        if isinstance(color, float):
            color = format(int(color), '06X')
        color = openpyxl.styles.colors.Color(rgb=color)
    cell.font = openpyxl.styles.Font(color=color)

def applyFontNameToCell(cell, name):
    if name is None or pd.isna(name):
        return
    cell.font = openpyxl.styles.Font(name=name)

def applyStyleToCell(cell, style):
    if style is None or pd.isna(style):
        return
    # check if the style is defined in STYLE_DEFINITIONS
    if style in STYLE_DEFINITIONS:
        if 'fill' in STYLE_DEFINITIONS[style]:
            cell.fill = STYLE_DEFINITIONS[style]['fill']
        if 'font' in STYLE_DEFINITIONS[style]:
            cell.font = STYLE_DEFINITIONS[style]['font']
        if 'border' in STYLE_DEFINITIONS[style]:
            cell.border = STYLE_DEFINITIONS[style]['border']
    else:
        print(f"Style {style} not found in STYLE_DEFINITIONS")

def apply_conditional_formatting(ws, report_columns, start_row=2, debug=False):
    for idx, col in enumerate(report_columns, start=1):
        col_letter = ws.cell(row=1, column=idx).column_letter
        cf_type = col.get('ConditionalFormatting')

        if not cf_type:
            continue  # Skip if no conditional formatting needed

        last_row = ws.max_row
        range_ref = f"{col_letter}{start_row}:{col_letter}{last_row}"  # all rows from start_row to last row

        if cf_type == 'ZeroOne':
            if debug:
                print(f"ZeroOne: {range_ref}")
            # Zero should have no formatting
            ws.conditional_formatting.add(range_ref,
                CellIsRule(operator='equal', formula=['0'], fill=PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')))
            # One should have a light orange background with dark orange text   
            ws.conditional_formatting.add(range_ref,
                CellIsRule(operator='equal', formula=['1'], fill=PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')))

        elif cf_type == 'ZeroTwo':
            if debug:
                print(f"ZeroTwo: {range_ref}")
            # Zero should have a light red background with dark red text
            ws.conditional_formatting.add(range_ref,
                CellIsRule(operator='equal', formula=['0'], fill=PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')))
            # Two should have a light green background with dark green text
            ws.conditional_formatting.add(range_ref,
                CellIsRule(operator='equal', formula=['2'], fill=PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid')))

        elif cf_type == 'TrueFalse':
            if debug:
                print(f"TrueFalse: {range_ref}")
            # True should have a light green background with dark green text
            ws.conditional_formatting.add(range_ref,
                CellIsRule(operator='equal', formula=['TRUE'], fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')))
            # False should have a light red background with dark red text
            ws.conditional_formatting.add(range_ref,
                CellIsRule(operator='equal', formula=['FALSE'], fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')))

        elif cf_type == 'GoodBadNA':
            if debug:
                print(f"GoodBadNA: {range_ref}")
            # To start set a 2 px white border around the cell
            ws.conditional_formatting.add(range_ref,
                CellIsRule(operator='equal', formula=['""'], fill=PatternFill(start_color='FFFFFF', end_color='FFFFFF', 
                                                                fill_type='solid'), border=Border(
                                                                        left=Side(style='thin', color='000000'), 
                                                                        right=Side(style='thin', color='000000'), 
                                                                        top=Side(style='thin', color='000000'), 
                                                                        bottom=Side(style='thin', color='000000'))))

            # GOOD should have a green background  green text
            ws.conditional_formatting.add(range_ref,
                CellIsRule(operator='equal', formula=['1'], fill=PatternFill(start_color='63EF45', end_color='63EF45', fill_type='solid'), font=Font(color='63EF45')))
            # BAD should have a red background red text
            ws.conditional_formatting.add(range_ref,
                CellIsRule(operator='equal', formula=['0'], fill=PatternFill(start_color='C0504D', end_color='C0504D', fill_type='solid'), font=Font(color='C0504D')))
            # blank cells should have no formatting
            ws.conditional_formatting.add(range_ref,
                CellIsRule(operator='equal', formula=['""'], fill=PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'), font=Font(color='FFFFFF')))
            
        elif cf_type == 'XBlank':
            if debug:
                print(f"XBlank: {range_ref}")
            # X should have a light orange background with dark orange text
            ws.conditional_formatting.add(range_ref,
                FormulaRule(formula=[f'EXACT({col_letter}{start_row},"X")'], 
                            fill=PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')))
            # Blank should have no formatting
            ws.conditional_formatting.add(range_ref,
                FormulaRule(formula=[f'ISBLANK({col_letter}{start_row})'], 
                            fill=PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')))

        elif cf_type == 'Bold':
            if debug:
                print(f"Bold: {range_ref}")
            ws.conditional_formatting.add(range_ref,
                FormulaRule(formula=[f'LEN(TRIM({col_letter}{start_row}))>0'], 
                            font=Font(bold=True)))

        elif cf_type == 'Italic':
            if debug:
                print(f"Italic: {range_ref}")
            ws.conditional_formatting.add(range_ref,
                FormulaRule(formula=[f'LEN(TRIM({col_letter}{start_row}))>0'], 
                            font=Font(italic=True)))

def create_points_section(df: pd.DataFrame) -> pd.DataFrame:
    """Create the points section of the report."""
    points = []
    
    for _, row in df.iterrows():
        
        if row.get('GenericType') in ['SD', 'DD']:
            point = {
                'Type': row.get('GenericType', ''),
                'SCADA Address': row.get('GenericPointAddress', ''),
                'eTerra Key': row.get('eTerraKey', ''),
                'PowerOn Alias': row.get('POAlias', ''),
                'ICCP Flag': row.get('ICCPFlag', ''),
                'Habdde Match Status': row.get('HbddeCompareStatus', ''),
                'PowerOn Config Health Status': row.get('ConfigHealth', ''),
                'Control Zone Status': row.get('CompAlarmAlarmZoneMatch', ''),
            }
            # Get the Ctrl Info
            if row.get('Controllable') == '1':
                if row.get('Ctrl1Addr','') != '':
                    point['Ctrl1Addr'] = row.get('Ctrl1Addr','')
                    point['Ctrl1Name'] = row.get('Ctrl1Name','')
                if row.get('Ctrl2Addr','') != '':
                    point['Ctrl2Addr'] = row.get('Ctrl2Addr','')
                    point['Ctrl2Name'] = row.get('Ctrl2Name','')
            else:
                point['Ctrl1Addr'] = ''
                point['Ctrl1Name'] = ''
                point['Ctrl2Addr'] = ''
                point['Ctrl2Name'] = ''

            # Get the Alarm Info
            if row.get('CompAlarmEterraAlias','') != '':
                point['CompAlarmeTerraAlarmZone'] = row.get('CompAlarmeTerraAlarmZone','')
                point['CompAlarmeTerraStatus'] = row.get('CompAlarmeTerraStatus','')
                point['CompAlarmPOsubstation'] = row.get('CompAlarmPOsubstation','')
                point['CompAlarmPOAlarmZone'] = row.get('CompAlarmPOAlarmZone','')
                point['CompAlarmPOAlarmRef'] = row.get('CompAlarmPOAlarmRef','')
                point['CompAlarmPOStatus'] = row.get('CompAlarmPOStatus','')
                point['CompAlarmAlarmZoneMatch'] = row.get('CompAlarmAlarmZoneMatch','')
                point['Alarm0_MessageMatch'] = row.get('Alarm0_MessageMatch','')
                point['Alarm1_MessageMatch'] = row.get('Alarm1_MessageMatch','')
                point['Alarm2_MessageMatch'] = row.get('Alarm2_MessageMatch','')
                point['Alarm3_MessageMatch'] = row.get('Alarm3_MessageMatch','')

            # Add the Report flags
            point['Report1'] = row.get('Report1', '')
            point['Report2'] = row.get('Report2', '')
            point['Report3'] = row.get('Report3', '')


            points.append(point)

    return pd.DataFrame(points)


def save_reports(reports: list, output_path: Path):
    """Save the report to an Excel file."""
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for report in reports:
            df = report['Content']
            df.to_excel(writer, sheet_name=report['RTU'], index=False)
        
        # Apply formatting
        worksheet = writer.sheets[report['RTU']]
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width 


def generate_report_in_excel(df: pd.DataFrame, report_definition: dict, output_path: Path):
    """Generate a report in Excel."""
    print(f"  🧠 Generating report {report_definition['name']} in excel ...")
    report_name = report_definition['name']
    report_worksheet_name = report_definition['worksheet_name'] if 'worksheet_name' in report_definition else 'Sheet1'
    report_columns = report_definition['columns']

    for idx, col in enumerate(report_columns, 1):
        if col['dfCol'] not in df.columns:
            print(f"Column {col['dfCol']} not found in df ... adding empty column")
            df[col['dfCol']] = ''  # Add empty column if missing
        # if the ColName is different from the dfCol, then rename the column
        if col['ColName'] != col['dfCol']:
            df = df.rename(columns={col['dfCol']: col['ColName']})

    report_fields = [col['ColName'] for col in report_columns]

    # create a new dataframe with the report fields and the new columns
    report_df = pd.DataFrame(columns=report_fields)

    # First ensure all required columns exist in merged_data
    for col in report_fields:
        if col not in df.columns:
            print(f"Column {col} not found in df ... adding empty column")
            df[col] = ''  # Add empty column if missing
            
    # Now we can safely select and concat

    # Create a single DataFrame with all fields, avoiding empty/NA concatenation
    report_df = df[report_fields].copy()

    # save the report dataframe to an xlsx file with formatting
    writer = pd.ExcelWriter(output_path / f"{report_name}.xlsx", engine='openpyxl')
    # Convert '1' and '0' to numbers but keep empty values as empty
    for col in report_df.columns:
        report_df[col] = report_df[col].apply(lambda x: int(x) if str(x) in ['0','1'] else x)
    report_df.to_excel(writer, sheet_name=report_worksheet_name, index=False, na_rep='')
    # Get the worksheet
    worksheet = writer.sheets[report_worksheet_name]
    # Add filters to row 1
    worksheet.auto_filter.ref = worksheet.dimensions
    
    # Format header row
    for cell in worksheet[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')
        
    # Freeze top row
    worksheet.freeze_panes = worksheet['F2']

    for idx, col in enumerate(report_columns, 1):

        # Rename report columns to be more readable going from dfCol to ColName 
        if col['dfCol'] != col['ColName']:
            worksheet[f"{get_column_letter(idx)}1"].value = col['ColName']

        # Setup column widths
        worksheet.column_dimensions[get_column_letter(idx)].width = col['ColWidth']

        # apply the alignment
        for row in range(1, len(report_df) + 2):
            worksheet[f"{get_column_letter(idx)}{row}"].alignment = openpyxl.styles.Alignment(horizontal=col['Align'])

        # Apply the fill colors and set all borders to be black and 1pt thick
        for row in range(2, len(report_df) + 2):
                cell = worksheet[f"{get_column_letter(idx)}{row}"]
                if 'ColFill' in col and col['ColFill']:
                    if isinstance(col['ColFill'], str):
                        cell.fill = openpyxl.styles.PatternFill(start_color=col['ColFill'], end_color=col['ColFill'], fill_type='solid')
                if 'FontStyle' in col and col['FontStyle']: # if a formatting style is specified, apply it
                    applyFontStyleToCell(cell, col['FontStyle'])
                if 'FontSize' in col and col['FontSize']:
                    applyFontSizeToCell(cell, col['FontSize'])
                if 'FontColor' in col and col['FontColor']:
                    applyFontColorToCell(cell, col['FontColor'])
                if 'FontName' in col and col['FontName']:
                    applyFontNameToCell(cell, col['FontName'])
                if 'Style' in col and col['Style']:
                    applyStyleToCell(cell, col['Style'])

                cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin', color='000000'),
                                                    right=openpyxl.styles.Side(style='thin', color='000000'),
                                                    top=openpyxl.styles.Side(style='thin', color='000000'),
                                                    bottom=openpyxl.styles.Side(style='thin', color='000000'))

    apply_conditional_formatting(worksheet, report_columns)

    # Hide the columns that are hidden
    for idx, col in enumerate(report_columns, 1):
        if 'Hidden' in col and col['Hidden']:
            worksheet.column_dimensions[get_column_letter(idx)].hidden = True

    writer.close()
    
    print(f"Defect report generated successfully: {output_path / f'{report_name}.xlsx'}")


def generate_defect_report_in_excel(df: pd.DataFrame, output_path: Path):

    print("="*80)
    print("Generating Defect Report in Excel")
    print("="*80)

    report_columns = [
        {'dfCol': 'GenericPointAddress',            'ColName': 'GenericPointAddress',           'ColWidth': 25,     'Align': 'left',    'ColFill': None,        'ConditionalFormatting': None,      'Hidden': False},
        {'dfCol': 'Type',                           'ColName': 'Type',                          'ColWidth': 3,      'Align': 'center',  'ColFill': None,        'ConditionalFormatting': None,      'Hidden': False},
        {'dfCol': 'RTU',                            'ColName': 'RTU',                           'ColWidth': 7,      'Align': 'center',  'ColFill': None,        'ConditionalFormatting': None,      'Hidden': False}, 
        {'dfCol': 'Sub',                            'ColName': 'Sub',                           'ColWidth': 7,      'Align': 'center',  'ColFill': None,        'ConditionalFormatting': None,      'Hidden': False},
        {'dfCol': 'Ignore',                         'ColName': 'Ignore',                        'ColWidth': 7,      'Align': 'center',  'ColFill': None,        'ConditionalFormatting': None,      'Hidden': False},
        {'dfCol': 'eTerraKey',                      'ColName': 'eTerraKey',                     'ColWidth': 17,     'Align': 'left',    'ColFill': None,        'ConditionalFormatting': None,      'Hidden': False},
        {'dfCol': 'eTerraAlias',                    'ColName': 'eTerraAlias',                   'ColWidth': 35,     'Align': 'left',    'ColFill': None,        'ConditionalFormatting': None,      'Hidden': False},
        {'dfCol': 'eTerraAliasExistsInPO',          'ColName': 'eTerraAliasExistsInPO',         'ColWidth': 2,      'Align': 'center',  'ColFill': None,        'ConditionalFormatting': 'GoodBadNA',      'Hidden': False},
        {'dfCol': 'eTerraAliasLinkedToSCADA',       'ColName': 'eTerraAliasLinkedToSCADA',      'ColWidth': 2,      'Align': 'center',  'ColFill': None,        'ConditionalFormatting': 'GoodBadNA',      'Hidden': False},
        {'dfCol': 'GridIncomer',                    'ColName': 'GridIncomer',                   'ColWidth': 10,     'Align': 'center',  'ColFill': None,        'ConditionalFormatting': 'ZeroOne',      'Hidden': False},
        {'dfCol': 'RTUComms',                       'ColName': 'RTUComms',                      'ColWidth': 7,     'Align': 'center',  'ColFill': None,         'ConditionalFormatting': None,      'Hidden': False},
        {'dfCol': 'TopLocation',                    'ColName': 'TopLocation',                   'ColWidth': 10,     'Align': 'left',    'ColFill': None,        'ConditionalFormatting': None,      'Hidden': False},
        {'dfCol': 'PointId',                        'ColName': 'PointId',                       'ColWidth': 5,     'Align': 'center',  'ColFill': None,         'ConditionalFormatting': 'None',      'Hidden': False},
        {'dfCol': 'ICCP->PO',                       'ColName': 'ICCP->PO',                      'ColWidth': 7,      'Align': 'center',  'ColFill': None,        'ConditionalFormatting': 'XBlank',      'Hidden': False},
        {'dfCol': 'ICCP_ALIAS',                     'ColName': 'ICCP_ALIAS',                    'ColWidth': 27,     'Align': 'left',    'ColFill': None,        'ConditionalFormatting': None,      'Hidden': False},
        {'dfCol': 'ICCPAliasExists',                'ColName': 'ICCPAliasExists',               'ColWidth': 2,      'Align': 'center',  'ColFill': None,        'ConditionalFormatting': 'GoodBadNA',      'Hidden': False},
        {'dfCol': 'ICCPAliasLinkedToSCADA',         'ColName': 'ICCPAliasLinkedToSCADA',        'ColWidth': 2,      'Align': 'center',  'ColFill': None,        'ConditionalFormatting': 'GoodBadNA',      'Hidden': False},
        {'dfCol': 'PowerOn Alias',                  'ColName': 'PowerOn Alias',                 'ColWidth': 35,     'Align': 'left',    'ColFill': None,        'ConditionalFormatting': None,      'Hidden': False},
        {'dfCol': 'CompAlarmTemplateAlias',         'ColName': 'Template'                  ,    'ColWidth': 20,      'Align': 'center',  'ColFill': None,        'ConditionalFormatting': 'None',      'Hidden': False},
        {'dfCol': 'PowerOn Alias Exists',           'ColName': 'PowerOn Alias Exists',          'ColWidth': 8,      'Align': 'center',  'ColFill': None,        'ConditionalFormatting': 'TrueFalse',      'Hidden': False},
        {'dfCol': 'PowerOn Alias Linked to SCADA',  'ColName': 'PowerOn Alias Linked to SCADA', 'ColWidth': 2,      'Align': 'center',  'ColFill': None,        'ConditionalFormatting': 'ZeroTwo',      'Hidden': False},
        {'dfCol': 'CompAlarmTemplateType',          'ColName': 'TemplateType',                  'ColWidth': 3,     'Align': 'center',  'ColFill': None,        'ConditionalFormatting': 'None',      'Hidden': False},
        {'dfCol': 'CompAlarmStateIndex',            'ColName': 'StateIndex',                    'ColWidth': 3,     'Align': 'center',  'ColFill': None,        'ConditionalFormatting': 'None',      'Hidden': False},
        {'dfCol': 'Alarm0',                         'ColName': 'Alarm0',                        'ColWidth': 4,      'Align': 'center',  'ColFill': None,        'ConditionalFormatting': 'GoodBadNA',      'Hidden': False},
        {'dfCol': 'Alarm1',                         'ColName': 'Alarm1',                        'ColWidth': 4,      'Align': 'center',  'ColFill': None,        'ConditionalFormatting': 'GoodBadNA',      'Hidden': False},
        {'dfCol': 'Alarm2',                         'ColName': 'Alarm2',                        'ColWidth': 4,      'Align': 'center',  'ColFill': None,        'ConditionalFormatting': 'GoodBadNA',      'Hidden': False},
        {'dfCol': 'Alarm3',                         'ColName': 'Alarm3',                        'ColWidth': 4,      'Align': 'center',  'ColFill': None,        'ConditionalFormatting': 'GoodBadNA',      'Hidden': False},
        {'dfCol': 'Alarm0_eTerraMessage',            'ColName': 'Alarm0_eTerraMessage',         'ColWidth': 20,     'Align': 'left',  'ColFill': None,        'ConditionalFormatting': 'Italic',      'Hidden': True},
        {'dfCol': 'Alarm0_POMessage',                'ColName': 'Alarm0_POMessage',             'ColWidth': 20,     'Align': 'left',  'ColFill': None,        'ConditionalFormatting': 'Italic',      'Hidden': True},
        {'dfCol': 'Alarm1_eTerraMessage',            'ColName': 'Alarm1_eTerraMessage',         'ColWidth': 20,     'Align': 'left',  'ColFill': None,        'ConditionalFormatting': 'Italic',      'Hidden': True},
        {'dfCol': 'Alarm1_POMessage',                'ColName': 'Alarm1_POMessage',             'ColWidth': 20,     'Align': 'left',  'ColFill': None,        'ConditionalFormatting': 'Italic',      'Hidden': True},
        {'dfCol': 'Alarm2_eTerraMessage',            'ColName': 'Alarm2_eTerraMessage',         'ColWidth': 20,     'Align': 'left',  'ColFill': None,        'ConditionalFormatting': 'Italic',      'Hidden': True},
        {'dfCol': 'Alarm2_POMessage',                'ColName': 'Alarm2_POMessage',             'ColWidth': 20,     'Align': 'left',  'ColFill': None,        'ConditionalFormatting': 'Italic',      'Hidden': True},
        {'dfCol': 'Alarm3_eTerraMessage',            'ColName': 'Alarm3_eTerraMessage',         'ColWidth': 20,     'Align': 'left',  'ColFill': None,        'ConditionalFormatting': 'Italic',      'Hidden': True},
        {'dfCol': 'Alarm3_POMessage',                'ColName': 'Alarm3_POMessage',             'ColWidth': 20,     'Align': 'left',  'ColFill': None,        'ConditionalFormatting': 'Italic',      'Hidden': True},
        {'dfCol': 'Controllable',                   'ColName': 'Controllable',                  'ColWidth': 10,     'Align': 'center',  'ColFill': None,        'ConditionalFormatting': None,      'Hidden': False},
        {'dfCol': 'Ctrl1',                          'ColName': 'Ctrl1',                         'ColWidth': 4,      'Align': 'center',  'ColFill': None,        'ConditionalFormatting': 'GoodBadNA',      'Hidden': False},
        {'dfCol': 'Ctrl1V',                         'ColName': 'Ctrl1V',                        'ColWidth': 1,      'Align': 'center',  'ColFill': None,        'ConditionalFormatting': 'GoodBadNA',      'Hidden': True},
        {'dfCol': 'Ctrl1C',                         'ColName': 'Ctrl1C',                        'ColWidth': 1,      'Align': 'center',  'ColFill': None,        'ConditionalFormatting': 'GoodBadNA',      'Hidden': True},
        {'dfCol': 'Ctrl2',                          'ColName': 'Ctrl2',                         'ColWidth': 4,      'Align': 'center',  'ColFill': None,        'ConditionalFormatting': 'GoodBadNA',      'Hidden': False},
        {'dfCol': 'Ctrl2V',                         'ColName': 'Ctrl2V',                        'ColWidth': 1,      'Align': 'center',  'ColFill': None,        'ConditionalFormatting': 'GoodBadNA',      'Hidden': True},
        {'dfCol': 'Ctrl2C',                         'ColName': 'Ctrl2C',                        'ColWidth': 1,      'Align': 'center',  'ColFill': None,        'ConditionalFormatting': 'GoodBadNA',      'Hidden': True},
        {'dfCol': 'Ctrl1Name',                      'ColName': 'Ctrl1Name',                     'ColWidth': 10,     'Align': 'left',    'ColFill': None,        'ConditionalFormatting': 'Bold',      'Hidden': False},
        {'dfCol': 'Ctrl1Comments',                  'ColName': 'Ctrl1Comments',                 'ColWidth': 10,     'Align': 'left',    'ColFill': None,        'ConditionalFormatting': 'Italic',      'Hidden': True},
        {'dfCol': 'Ctrl1ConfigHealth',              'ColName': 'Ctrl1ConfigHealth',             'ColWidth': 10,     'Align': 'left',    'ColFill': None,        'ConditionalFormatting': 'Italic',      'Hidden': False},
        {'dfCol': 'Ctrl2Name',                      'ColName': 'Ctrl2Name',                     'ColWidth': 10,     'Align': 'left',    'ColFill': None,        'ConditionalFormatting': 'Bold',      'Hidden': False},
        {'dfCol': 'Ctrl2Comments',                  'ColName': 'Ctrl2Comments',                 'ColWidth': 10,     'Align': 'left',    'ColFill': None,        'ConditionalFormatting': 'Italic',      'Hidden': True},
        {'dfCol': 'Ctrl2ConfigHealth',              'ColName': 'Ctrl2ConfigHealth',             'ColWidth': 10,     'Align': 'left',    'ColFill': None,        'ConditionalFormatting': 'Italic',      'Hidden': False},
        {'dfCol': 'AlarmMismatchComment',           'ColName': 'AlarmMismatchComment',          'ColWidth': 10,     'Align': 'left',    'ColFill': None,        'ConditionalFormatting': None,      'Hidden': False},
        {'dfCol': 'AlarmMismatchTemplateAlias',     'ColName': 'AlarmMismatchTemplateAlias',    'ColWidth': 10,     'Align': 'left',    'ColFill': None,        'ConditionalFormatting': None,      'Hidden': False},
        {'dfCol': 'T3 Analysis',                    'ColName': 'T3 Analysis',                   'ColWidth': 20,     'Align': 'left',    'ColFill': None,        'ConditionalFormatting': None,      'Hidden': False},
        {'dfCol': 'T5 Analysis',                    'ColName': 'T5 Analysis',                   'ColWidth': 20,     'Align': 'left',    'ColFill': None,        'ConditionalFormatting': None,      'Hidden': False},
        {'dfCol': 'Report1',            'ColName': 'Missing Analog Components',                 'ColWidth': 8,      'Align': 'center',  'ColFill': None,        'ConditionalFormatting': 'TrueFalse',      'Hidden': False},
        {'dfCol': 'Report2',            'ColName': 'Missing Digital Components',                'ColWidth': 8,      'Align': 'center',  'ColFill': None,        'ConditionalFormatting': 'TrueFalse',      'Hidden': False},
        {'dfCol': 'Report3',            'ColName': 'Missing Controllable Components',           'ColWidth': 8,      'Align': 'center',  'ColFill': None,        'ConditionalFormatting': 'TrueFalse',      'Hidden': False},
        {'dfCol': 'Report4',            'ColName': 'Components Missing Telecontrol Actions',    'ColWidth': 8,      'Align': 'center',  'ColFill': None,        'ConditionalFormatting': 'TrueFalse',      'Hidden': False},
        {'dfCol': 'Report5',            'ColName': 'Components Missing Alarm Reference',        'ColWidth': 8,      'Align': 'center',  'ColFill': None,        'ConditionalFormatting': 'TrueFalse',      'Hidden': False},
        {'dfCol': 'Report6',            'ColName': 'Controls not in PO but tested ok',          'ColWidth': 8,      'Align': 'center',  'ColFill': None,        'ConditionalFormatting': 'TrueFalse',      'Hidden': False},
        {'dfCol': 'Report7',            'ColName': 'Controls Not Linked',                       'ColWidth': 8,      'Align': 'center',  'ColFill': None,        'ConditionalFormatting': 'TrueFalse',      'Hidden': False},
        {'dfCol': 'Report8',            'ColName': 'Ctrl-able eTerra Points with no Controls',  'ColWidth': 8,      'Align': 'center',  'ColFill': None,        'ConditionalFormatting': 'TrueFalse',      'Hidden': False},
        {'dfCol': 'Report9',            'ColName': 'Alarm Mismatch Manual Actions',             'ColWidth': 8,      'Align': 'center',  'ColFill': None,        'ConditionalFormatting': 'TrueFalse',      'Hidden': False},
        {'dfCol': 'Report13',           'ColName': 'SD symbol should be DD',                    'ColWidth': 8,      'Align': 'center',  'ColFill': None,        'ConditionalFormatting': 'TrueFalse',      'Hidden': False},
        {'dfCol': 'Report14',           'ColName': 'DD symbol should be SD',                    'ColWidth': 8,      'Align': 'center',  'ColFill': None,        'ConditionalFormatting': 'TrueFalse',      'Hidden': False},
        {'dfCol': 'ReportANY',          'ColName': 'Any Defect',                                'ColWidth': 8,      'Align': 'center',  'ColFill': None,        'ConditionalFormatting': 'TrueFalse',      'Hidden': False},
        {'dfCol': 'Report10',           'ColName': 'RESET w/ CtrlFunc 0',                       'ColWidth': 8,      'Align': 'center',  'ColFill': None,        'ConditionalFormatting': 'TrueFalse',      'Hidden': True},
        {'dfCol': 'Report11',           'ColName': 'SWDD with LAMP symbol',                     'ColWidth': 8,      'Align': 'center',  'ColFill': None,        'ConditionalFormatting': 'TrueFalse',      'Hidden': False},
        {'dfCol': 'Report12',           'ColName': 'Missing from DLPoint',                      'ColWidth': 8,      'Align': 'center',  'ColFill': None,        'ConditionalFormatting': 'TrueFalse',      'Hidden': False},
        {'dfCol': 'Report15',           'ColName': 'ICCP SD Inverted but needs un-inverted',    'ColWidth': 8,      'Align': 'center',  'ColFill': None,        'ConditionalFormatting': 'TrueFalse',      'Hidden': True},
        {'dfCol': 'Report16',           'ColName': 'ICCP SD Inverted but in SPT hierarchy',     'ColWidth': 8,      'Align': 'center',  'ColFill': None,        'ConditionalFormatting': 'TrueFalse',      'Hidden': True},
        {'dfCol': 'Report17',           'ColName': '2 copies of COMP - ICCP is linked',         'ColWidth': 8,      'Align': 'center',  'ColFill': None,        'ConditionalFormatting': 'TrueFalse',      'Hidden': False},
        {'dfCol': 'Review Status',                  'ColName': 'Review Status',                 'ColWidth': 12,     'Align': 'left',    'ColFill': 'FFFFE0',    'ConditionalFormatting': None,      'Hidden': False},
        {'dfCol': 'Comments',                       'ColName': 'Comments',                      'ColWidth': 60,     'Align': 'left',    'ColFill': 'FFFFE0',    'ConditionalFormatting': None,      'Hidden': False}
    ]

    report_definition = {
        'name': 'defect_report_orig',
        'worksheet_name': 'Sheet1',
        'columns': report_columns
    }
    generate_report_in_excel(df, report_definition, output_path)

    return 

