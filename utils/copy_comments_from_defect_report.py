# We will read in  defect_report_all.xlsx and copy the coments to a new version of the same report
# we need the filenames to be explicity passed in as arguments
# we need to count how many are in the old one, check theres none in the new onw and ask teh user to confrim before making the update

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
from rich.progress import Progress
import copy
default_sheet_name = 'Sheet1'

ColumnsToCopy = ['Review Status', 'Comments', 'Assigned To']

def debug_a_row_in_wb(wb, ColumnName, MatchValue, sheet_name=""):
    if sheet_name == "":
        sheet_name = default_sheet_name
    ws = wb[sheet_name]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    column_index = header_row.index(ColumnName)  # Remove trailing comma
    
    for row in ws.iter_rows(min_row=2, values_only=False):
        cell_value = row[column_index].value if row[column_index] else None
        if cell_value == MatchValue:
            print(f"DEBUG> Found matching row:")
            for i, cell in enumerate(row):
                #print the value and the fill color of columns in ColumnsToCopy
                for print_col in ColumnsToCopy:
                    if header_row[i] == print_col:
                        if cell.fill and cell.fill.fill_type:
                            cell_fill_type = cell.fill.fill_type
                            cell_fill_fg_color = cell.fill.fgColor.rgb
                            cell_fill_bg_color = cell.fill.bgColor.rgb
                        else:
                            cell_fill_type = 'No fill'
                            cell_fill_fg_color = 'No fill'
                            cell_fill_bg_color = 'No fill'
                    
                        print(f"{MatchValue}: {header_row[i]}: {print_col}> {cell.value} <fill_type> {cell_fill_type} <fill_fg_color> {cell_fill_fg_color} <fill_bg_color> {cell_fill_bg_color} <")
            break


def read_report_df_and_wb(filename, sheet_name=default_sheet_name):
    # we want to read the values and fill colors for the following fields:
    # GenericPointAddress
    # eTerraAlias
    # + ColumnsToCopy

    print(f" :mag_right: Reading {filename} into a dataframe...")
    df= pd.read_excel(filename)
    df = df[['GenericPointAddress', 'eTerraAlias'] + ColumnsToCopy]

    print(f" :mag_right: Reading {filename} into a workbook...")
    # Load workbook and sheet using openpyxl
    wb = load_workbook(filename, data_only=True)

    return df, wb

def read_report_wb(filename):
    return load_workbook(filename, data_only=True)

def count_values_in_dict(wb_dict, value_to_count):
    count = 0
    for key in wb_dict:
        if wb_dict[key][value_to_count][0] is not None and wb_dict[key][value_to_count][0] != '':
            #print(f"DEBUG> Counting {value_to_count}: {wb_dict[key][value_to_count][0]}")
            count += 1
    return count

def count_values_in_df(df, value_to_count):
    return df[value_to_count].count()


def get_dict_of_values_and_fill_color(wb, matchmethod, sheet_name=""):
    # Create dictionaries to store the data in the worksheetfor fast lookup
    if sheet_name == "":
        sheet_name = default_sheet_name
    ws = wb[sheet_name]
    data = {}
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    matchmethod_index = header_row.index(matchmethod)
    column_indices = {col: header_row.index(col) for col in ColumnsToCopy}

    for row in ws.iter_rows(min_row=2, values_only=False):
        key = row[matchmethod_index].value
        if any(row[column_indices[col]].value for col in ColumnsToCopy):
            data[key] = {col: (row[column_indices[col]].value, row[column_indices[col]].fill) for col in ColumnsToCopy}
    
    return data


def copy_values_and_fill_color(old_wb, new_wb, matchmethod, old_sheet_name="", new_sheet_name=""):
    if old_sheet_name == "":
        old_sheet_name = default_sheet_name
    if new_sheet_name == "":
        new_sheet_name = default_sheet_name

    old_ws = old_wb[old_sheet_name]
    new_ws = new_wb[new_sheet_name]

    # Create dictionaries to store the old data for fast lookup
    old_data = {}
    old_header_row = next(old_ws.iter_rows(min_row=1, max_row=1, values_only=True))
    matchmethod_index_old = old_header_row.index(matchmethod)
    column_indices = {col: old_header_row.index(col) for col in ColumnsToCopy}

    # Build lookup dictionary from old worksheet - only store non-empty cells
    print("Building lookup dictionary from old worksheet...")
    for row in old_ws.iter_rows(min_row=2, values_only=False):
        key = row[matchmethod_index_old].value
        if any(row[column_indices[col]].value for col in ColumnsToCopy):
            old_data[key] = {col: (row[column_indices[col]].value, row[column_indices[col]].fill) for col in ColumnsToCopy}

    # Get new worksheet structure
    new_header_row = next(new_ws.iter_rows(min_row=1, max_row=1, values_only=True))
    matchmethod_index_new = new_header_row.index(matchmethod)
    new_column_indices = {col: new_header_row.index(col) + 1 for col in ColumnsToCopy}

    # Process new worksheet using the lookup dictionary
    print("Copying data to new worksheet...")
    with Progress() as progress:
        task = progress.add_task("Processing rows...", total=len(old_data))
        
        for row in new_ws.iter_rows(min_row=2, values_only=False):
            key = row[matchmethod_index_new].value
            if key in old_data:
                for col_name in ColumnsToCopy:
                    old_value, old_fill = old_data[key][col_name]
                    new_cell = row[new_column_indices[col_name] - 1]
                    
                    if new_cell.value != old_value:
                        new_cell.value = old_value
                        if old_fill is not None:
                            new_cell.fill = copy.copy(old_fill)
                progress.advance(task)


def get_params():
    import argparse
    # we need the filenames to be explicity passed in as arguments

    parser = argparse.ArgumentParser(description=f"Copy columns {ColumnsToCopy} from defect report, values and fill color")
    parser.add_argument("--oldfile", required=True, help="Path to the old defect report file")
    parser.add_argument("--newfile", required=True, help="Path to the new defect report file")
    parser.add_argument("--oldsheetname", required=False, help="Sheet name to use in the old file")
    parser.add_argument("--newsheetname", required=False, help="Sheet name to use in the new file")
    parser.add_argument("--matchmethod", choices=["eTerraAlias", "GenericPointAddress"], default="GenericPointAddress", help="Method to match the rows between the two files")
    args = parser.parse_args()

    # Check if the files exist
    if not os.path.exists(args.oldfile):
        print(f"Error: The file {args.oldfile} does not exist.")
        exit(1)
    if not os.path.exists(args.newfile):
        print(f"Error: The file {args.newfile} does not exist.")
        exit(1)

    if args.oldsheetname is None:
        args.oldsheetname = default_sheet_name
    if args.newsheetname is None:
        args.newsheetname = default_sheet_name

    print("")
    print(f"Using matchmethod: {args.matchmethod} to copy columns {ColumnsToCopy} from {args.oldfile} sheet {args.oldsheetname} to {args.newfile} sheet {args.newsheetname}\n")
    return args.oldfile, args.newfile, args.matchmethod, args.oldsheetname, args.newsheetname

def main():
    old_file, new_file, matchmethod, old_sheet_name, new_sheet_name = get_params()

    # we need to count how many are in the old one, check theres none in the new onw and ask teh user to confrim before making the update
    old_wb = read_report_wb(old_file)
    # debug_a_row_in_wb(old_wb, "eTerraAlias", "TONG1/011_CB/661_13/AMPS")
    # debug_a_row_in_wb(old_wb, "eTerraAlias", "NOKY1/033_SC/WF_B/MW")
    old_dict = get_dict_of_values_and_fill_color(old_wb, matchmethod, old_sheet_name)
    new_wb = read_report_wb(new_file)
    new_dict = get_dict_of_values_and_fill_color(new_wb, matchmethod, new_sheet_name)

    print(f"Opened old file {old_file} and new file {new_file}...")

    for col in ColumnsToCopy:
        old_count = count_values_in_dict(old_dict, col)
        new_count = count_values_in_dict(new_dict, col)
        print(f"Old {col} count: {old_count}")
        print(f"New {col} count: {new_count}")

    # Ask the user to confirm before making the update
    user_confirm = input("Are you sure you want to make the update? (y/n): ")
    if user_confirm != "y":
        print("Update cancelled.")
        return

    # Make the update for the values and fill color
    copy_values_and_fill_color(old_wb, new_wb, matchmethod, old_sheet_name, new_sheet_name)

    # Save the new file
    new_wb.save(new_file)

    print(f"Updated {new_file} with the values and fill color from {old_file}.")


if __name__ == "__main__":
    main() 

